import random
import tkinter as tk
from ttkbootstrap import Style
from tkinter.scrolledtext import ScrolledText
from ttkbootstrap.widgets import Entry, Label, Button, LabelFrame
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import threading
import time

executando = False
produtos = {
    "Produto 1": 100.0,
    "Produto 2": 200.0,
    "Produto 3": 300.0,
    "Produto 4": 400.0,
    "Produto 5": 500.0,
}

# Definições no escopo global
nome_entry = None
cliente_entry = None
cnpj_entry = None
intervalo_entry = None
quantidade_entry = None
erro_entry = None
icms_entry = None
ipi_entry = None
pis_entry = None
cofins_entry = None
produto_entries = []
preco_entries = []


def inicializar_excel(nome_arquivo):
    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"
    if not os.path.exists(nome_arquivo):
        wb = Workbook()
        ws = wb.active
        ws.title = "Notas"
        ws.append(
            [
                "Data",
                "Cliente",
                "CNPJ",
                "Produto",
                "Preço Unitário",
                "ICMS",
                "IPI",
                "PIS",
                "COFINS",
                "Valor Total",
            ]
        )
        wb.save(nome_arquivo)


def salvar_linha_excel(linha, nome_arquivo):
    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"

    if not os.path.exists(nome_arquivo):
        wb = Workbook()
        ws = wb.active
        ws.title = "Lançamentos"
        ws.append(
            [
                "Data",
                "Cliente",
                "CNPJ",
                "Produto",
                "Preço Unitário",
                "ICMS",
                "IPI",
                "PIS",
                "COFINS",
                "Valor Total",
            ]
        )
        wb.save(nome_arquivo)

    wb = load_workbook(nome_arquivo)
    ws = wb.active
    ws.append(linha)
    wb.save(nome_arquivo)


def erro_ocorre(probabilidade_percentual):
    return random.random() < (probabilidade_percentual / 100)


def aplicar_erro(valor):
    valor_str = str(valor)
    erro_tipo = random.choice(["virgula", "letra", "soma_errada", "numero_errado"])
    if erro_tipo == "virgula":
        return valor_str.replace(".", ",", 1)
    elif erro_tipo == "letra":
        pos = random.randint(0, len(valor_str) - 1)
        return (
            valor_str[:pos]
            + random.choice("abcdefghijklmnopqrstuvwxyz")
            + valor_str[pos:]
        )
    elif erro_tipo == "soma_errada":
        erro = random.uniform(-10, 10)
        return round(float(valor) + erro, 2)
    elif erro_tipo == "numero_errado":
        return valor_str[:-1] + str(random.randint(0, 9))
    return valor_str


def gerar_nota(cliente, cnpj, impostos, prob_erro):
    produto = random.choice(list(produtos.keys()))
    preco_unitario = produtos[produto]
    impostos_valores = {}
    for nome, porcentagem in impostos.items():
        imposto = preco_unitario * (porcentagem / 100)
        if erro_ocorre(prob_erro):
            imposto = aplicar_erro(imposto)
        else:
            imposto = round(imposto, 2)
        impostos_valores[nome] = imposto

    valor_total = preco_unitario
    for imposto in impostos_valores.values():
        try:
            valor_total += float(imposto)
        except ValueError:
            pass

    if erro_ocorre(prob_erro):
        preco_unitario = aplicar_erro(preco_unitario)

    return [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        cliente,
        cnpj,
        produto,
        preco_unitario,
        impostos_valores["ICMS"],
        impostos_valores["IPI"],
        impostos_valores["PIS"],
        impostos_valores["COFINS"],
        valor_total,
    ]


def loop_continuo(
    cliente,
    cnpj,
    impostos,
    intervalo,
    quantidade,
    nome_arquivo,
    status_label,
    prob_erro,
):
    global executando
    while executando:
        for _ in range(quantidade):
            nota = gerar_nota(cliente, cnpj, impostos, prob_erro)
            salvar_linha_excel(nota, nome_arquivo)
            mensagem = f"Nota lançada: {nota[3]} - R${nota[4]}"
            status_label.config(text=mensagem)
            escrever_log(mensagem)
        time.sleep(intervalo)

def nome_arquivo_valido(nome):
    nome = nome.strip().replace("/", "_").replace("\\", "_").replace(":", "_")
    if not nome.lower().endswith(".xlsx"):
        nome += ".xlsx"
    return nome


def iniciar():
    global executando, produtos

    global nome_entry, cliente_entry, cnpj_entry
    global intervalo_entry, quantidade_entry, erro_entry
    global icms_entry, ipi_entry, pis_entry, cofins_entry
    global produto_entries, preco_entries

    executando = True

    cliente = cliente_entry.get()
    cnpj = cnpj_entry.get()

    nome_arquivo = nome_entry.get().strip()

    if not nome_arquivo or nome_arquivo.lower().startswith("produto"):
        nome_arquivo = "notas_avancado.xlsx"

    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"


    intervalo = int(intervalo_entry.get())
    quantidade = int(quantidade_entry.get())
    prob_erro = float(erro_entry.get())

    impostos = {
        "ICMS": float(icms_entry.get()),
        "IPI": float(ipi_entry.get()),
        "PIS": float(pis_entry.get()),
        "COFINS": float(cofins_entry.get()),
    }

    produtos.clear()
    for nome_entry, preco_entry in zip(produto_entries, preco_entries):
        nome_produto = nome_entry.get().strip()
        if not nome_produto:
            continue  # pula produtos sem nome
        try:
            preco_produto = float(preco_entry.get())
            produtos[nome_produto] = preco_produto
        except ValueError:
            pass


    inicializar_excel(nome_arquivo)
    status_label.config(text="Bot rodando...", foreground="green")
    thread = threading.Thread(
        target=loop_continuo,
        args=(
            cliente,
            cnpj,
            impostos,
            intervalo,
            quantidade,
            nome_arquivo,
            status_label,
            prob_erro,
        ),
        daemon=True,
    )
    thread.start()


def parar():
    global executando
    executando = False
    status_label.config(text="Bot parado.", foreground="orange")


# GUI
root = tk.Tk()
root.title("Bot Lançador de NF Avançado")
style = Style("cyborg")
root.geometry("1200x800")

esquerda_frame = tk.Frame(root)
esquerda_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

direita_frame = tk.Frame(root)
direita_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)


def criar_label_entry(container, texto, entry_default=""):
    Label(container, text=texto).pack(anchor="w", padx=10, pady=(10, 0))
    entry = Entry(container)
    entry.insert(0, entry_default)
    entry.pack(fill="x", padx=10)
    return entry


# Lado Esquerdo
cliente_frame = LabelFrame(esquerda_frame, text="🧾 Dados do Cliente")
cliente_frame.pack(fill="x", pady=5)
cliente_entry = criar_label_entry(cliente_frame, "Cliente:", "Empresa Exemplo")
cnpj_entry = criar_label_entry(cliente_frame, "CNPJ:", "12.345.678/0001-90")

impostos_frame = LabelFrame(esquerda_frame, text="💰 Impostos (%)")
impostos_frame.pack(fill="x", pady=5)
icms_entry = criar_label_entry(impostos_frame, "ICMS (%):", "18")
ipi_entry = criar_label_entry(impostos_frame, "IPI (%):", "5")
pis_entry = criar_label_entry(impostos_frame, "PIS (%):", "1.65")
cofins_entry = criar_label_entry(impostos_frame, "COFINS (%):", "7.6")

config_frame = LabelFrame(esquerda_frame, text="⚙️ Configurações")
config_frame.pack(fill="x", pady=5)
nome_entry = criar_label_entry(
    config_frame, "Nome do Arquivo Excel:", "notas_avancado.xlsx"
)
intervalo_entry = criar_label_entry(
    config_frame, "Intervalo entre ciclos (segundos):", "5"
)
quantidade_entry = criar_label_entry(
    config_frame, "Quantidade de notas por ciclo:", "2"
)
erro_entry = criar_label_entry(config_frame, "Probabilidade de Erro (%):", "10")

# Produtos e Preços com adicionar/remover
produtos_frame = LabelFrame(direita_frame, text="📦 Produtos e Preços")
produtos_frame.pack(fill="both", expand=True, pady=5)
produto_entries = []
preco_entries = []


def atualizar_produtos_frame():
    for widget in produtos_frame.winfo_children():
        widget.destroy()

    # 🛑 Limpa as listas de entradas para evitar erro de referência
    produto_entries.clear()
    preco_entries.clear()

    for i, (nome, preco) in enumerate(produtos.items()):
        Label(produtos_frame, text=f"Produto {i+1}:").grid(
            row=i, column=0, sticky="w", padx=5, pady=5
        )

        produto_entry = Entry(produtos_frame, width=30)
        produto_entry.insert(0, nome)
        produto_entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")

        preco_entry = Entry(produtos_frame, width=10)
        preco_entry.insert(0, str(preco))
        preco_entry.grid(row=i, column=3, padx=5, pady=5, sticky="ew")

        produto_entries.append(produto_entry)
        preco_entries.append(preco_entry)

        Label(produtos_frame, text="Preço:").grid(
            row=i, column=2, sticky="w", padx=5, pady=5
        )

    botoes = tk.Frame(produtos_frame)
    botoes.grid(row=len(produtos), column=0, columnspan=4, pady=10)
    Button(
        botoes, text="➕ Adicionar Produto", bootstyle="info", command=adicionar_produto
    ).pack(side="left", padx=5)
    Button(
        botoes, text="➖ Remover Produto", bootstyle="danger", command=remover_produto
    ).pack(side="left", padx=5)

    produtos_frame.columnconfigure(1, weight=1)
    produtos_frame.columnconfigure(3, weight=1)


def adicionar_produto():
    novo_nome = f"Produto {len(produtos) + 1}"
    produtos[novo_nome] = 0.00
    atualizar_produtos_frame()


def remover_produto():
    if produtos:
        ultimo = list(produtos.keys())[-1]
        produtos.pop(ultimo)
        atualizar_produtos_frame()


# Inicializa com 5 produtos
for i in range(5):
    produto_entry = Entry(produtos_frame, width=30)
    produto_entry.insert(0, f"Produto {i+1}")
    preco_entry = Entry(produtos_frame, width=10)
    preco_entry.insert(0, str(produtos[f"Produto {i+1}"]))
    produto_entries.append(produto_entry)
    preco_entries.append(preco_entry)

atualizar_produtos_frame()

# Botões principais
botoes_frame = tk.Frame(root)
botoes_frame.pack(pady=15)
Button(
    botoes_frame, text="▶ Iniciar Lançamento", bootstyle="success", command=iniciar
).pack(side="left", padx=10)
Button(botoes_frame, text="■ Parar Lançamento", bootstyle="danger", command=parar).pack(
    side="left", padx=10
)

log_frame = LabelFrame(root, text="📜 Log de Execução")
log_frame.pack(fill="both", expand=True, padx=10, pady=5)

log_text = ScrolledText(log_frame, height=15, state="disabled")
log_text.pack(fill="both", expand=True, padx=5, pady=5)


def escrever_log(mensagem):
    log_text.config(state="normal")
    log_text.insert("end", f"{datetime.now().strftime('%H:%M:%S')} - {mensagem}\n")
    log_text.yview("end")  # Scroll automático para o final
    log_text.config(state="disabled")


status_label = Label(root, text="", anchor="center", font=("Segoe UI", 10, "bold"))
status_label.pack(pady=10)

root.mainloop()


config
